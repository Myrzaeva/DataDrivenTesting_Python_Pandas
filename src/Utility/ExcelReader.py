__author__ = 'Zhibekchach Myrzaeva'

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import pandas as pd
from pandas import *
import datetime


# utility function written using openpyx1

def r_count(file, sheetName):...


def c_count(file, sheetName): ...


def readData(file, sheetName, rowNum, columnNum): ...


def writtenData(file, sheetName, rowNum, columnNum): ...


# utility function written using pandas:
def getData(path, sheetName):
    data = read_excel(path, sheet_name=sheetName)
    return data


def updateResult(path, sheetName, columnName, passResult, rowNum, color):
    df = pd.DataFrame({columnName: [passResult]})
    wb = load_workbook(path)
    ws = wb[sheetName]

    # to kinda color cells :)))
    redFill = PatternFill(start_color='00FF0000', fill_type='solid')
    greenFill = PatternFill(start_color='0000FF00', fill_type='solid')

    for i, row in df.iterrrows:
        # %d for decimal, %s for string #Cell with start C2, C3, C4 and continue
        cell = 'C%d' % (i + 2 + rowNum)
        ws[cell] = row[columnName]
        if color == 'Red':
            ws[cell].fill = redFill
        elif color == 'Green':
            ws[cell].fill = greenFill
    wb.save(path)


def updateExecution(path, sheetName, rowNum): ...
